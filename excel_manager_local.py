"""
"from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from config import EXCEL_NAME
from cuotas_manager import estado_cuota


COLUMNAS = [
    "Fecha carga",
    "Pagador",
    "Alumno",
    "Monto base",
    "Recargo %",
    "Monto final",
    "Estado cuota",
    "Medio de pago",
    "Fecha comprobante",
    "Alias verificado",
    "Observaciones"
]


class ExcelManager:
    def __init__(self):
        if not os.path.exists(EXCEL_NAME):
            wb = Workbook()
            ws = wb.active
            ws.title = self._nombre_hoja_actual()
            ws.append(COLUMNAS)
            wb.save(EXCEL_NAME)

    def _nombre_hoja_actual(self):
        return datetime.now().strftime("%Y-%m")

    def _obtener_hoja_mes(self, wb):
        nombre = self._nombre_hoja_actual()
        if nombre not in wb.sheetnames:
            ws = wb.create_sheet(title=nombre)
            ws.append(COLUMNAS)
            return ws
        return wb[nombre]

    def registrar_pago(self, data: dict):
        wb = load_workbook(EXCEL_NAME)
        ws = self._obtener_hoja_mes(wb)

        hoy = datetime.now()
        cuota = estado_cuota(hoy)

        monto_base = float(data.get("monto_base", 0))
        monto_final = monto_base * cuota["factor"]

        ws.append([
            hoy.strftime("%d/%m/%Y %H:%M"),
            data.get("pagador", ""),
            data.get("alumno", ""),
            monto_base,
            f"{int(cuota['recargo'] * 100)}%",
            monto_final,
            cuota["estado"],
            data.get("medio", ""),
            data.get("fecha_comprobante", ""),
            data.get("alias_ok", ""),
            data.get("observaciones", "")
        ])

        wb.save(EXCEL_NAME)
"""